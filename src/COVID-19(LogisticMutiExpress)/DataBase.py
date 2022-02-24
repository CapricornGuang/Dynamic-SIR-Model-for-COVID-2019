from openpyxl.reader.excel import load_workbook
import numpy as np 

def Store(Data,wb=load_workbook(r'viruse.xlsx')):
    ws=wb.active
    if  isinstance(Data[0],list):
        for item in Data:
            ws.append(item)
    else:
        ws.append(Data)
    wb.save(r'viruse.xlsx')

def Extract(ws=load_workbook(r'viruse.xlsx').active):
    return np.array([ np.array([float(data.value) if data.value!=None else 0.39 for data in item]) for item in  ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=2,max_col=ws.max_column) ])
#为了保证数据的准确性，我们应当用 float数据类型 来计算！！None值处理为0.39
def EachProvinceData(province):
    flection={ '北京':0,'湖北':1,'广东':2,'浙江':3,'河南':4,'湖南':5,'重庆':6,'安徽':7,'四川':8,'山东':9,'吉林':10,'福建':11,'江西':12,'江苏':13,'上海':14,'广西':15,'海南':16,'陕西':17,'河北':18,'黑龙江':19,'辽宁':20,'云南':21,'天津':22,'山西':23,'甘肃':24,'内蒙古':25,'台湾':26,'澳门':27,'香港':28,'贵州':29,'西藏':30,'青海':31,'新疆':32,'宁夏':33}
    index=flection[province]
    completeData=Extract()
    res=completeData[:,index]
    data=np.insert(res.reshape(-1,1),0,np.arange(1,len(res)+1),axis=1)
    return data 
'''
for absent data ,it will be processed as None , but not nan 
print(Extract())
'''