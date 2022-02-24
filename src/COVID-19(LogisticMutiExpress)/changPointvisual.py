from openpyxl.reader.excel import load_workbook
from Analyse import DateList
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import numpy as np 
font=FontProperties(fname='res/simsun.ttc',size=20)
wb=load_workbook(r'viruse.xlsx')
ws=wb.active
stu_num=[]
for item in ws['C']:
    stu_num.append(item.value)
del(stu_num[0])
i,j=0,1
speed=list()
print(stu_num)
while(j<len(stu_num)):
    speed.append(stu_num[j]-stu_num[i])
    i,j=i+1,j+1
dateTime=list(DateList('2020-1-12',len(speed)))
xlabel=list(range(0,len(speed)))
#plt.scatter(xlabel,speed,s=5,c='red')
for num in range(len(speed)):
    if num%5==0:
        continue
    else:
        dateTime[num]=''
'''设置plot'''
xlabelPlot=xlabel
speedPlot=speed
for i in [32,31,24,23]:
    del(xlabelPlot[i])
    del(speedPlot[i])
#plt.scatter(xlabelPlot,speedPlot,s=5,c='red')
#plt.plot(xlabelPlot,speedPlot,c='red',linewidth=0.7)
plt.tick_params(labelsize=15)
plt.xticks(xlabel,dateTime)
plt.grid(linestyle=':')
plt.ylabel('每日新增病例(例)',FontProperties=font)
plt.xlabel('日期',FontProperties=font)
plt.show()