from openpyxl.reader.excel import load_workbook
import numpy as np
def Datasplit(m):
    data=np.array([62, 198, 198, 270, 375, 444, 549, 729, 1052, 1423, 2714, 3554, 4586, 5806, 7153, 9074, 11177, 13522, 13522, 19665, 22112, 24953, 27100, 29631, 31728, 33366, 33366, 48206, 51986, 54406, 56249, 58182, 59989, 61682, 62457, 63088, 63454, 63889, 64287, 64786, 65187, 65596, 65914, 66337, 66907, 67103, 67217, 67332, 67466, 67592, 67666, 67707, 67743, 67760, 67773, 67781, 67786, 67790, 67794, 67798, 67799, 67800, 67800, 67800, 67800, 67800, 67800, 67801, 67801, 67801, 67801, 67801, 67801, 67801, 67801, 67802, 67802, 67802])
    splitData=[data[m:6+m],data[6+m:]]
    return splitData
def readOut(wb=load_workbook(r'viruse.xlsx')):
    storeList=list()
    ws,i=wb.active,0
    for eachDay in ws['C']:
        if(i==0):
            i=1
            continue 
        else:
            eachDay=int(eachDay.value)
            storeList.append(eachDay)
    storeArray=np.array(storeList)
    return storeArray